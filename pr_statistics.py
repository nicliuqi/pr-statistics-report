import codecs
import csv
import datetime
import logging
import openpyxl
import os
import pandas as pd
import requests
import smtplib
import subprocess
import sys
import time
import yaml
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from logging import handlers
from openpyxl.styles import Alignment, Border, PatternFill, Side, Font
from xlsx2html import xlsx2html


class Logger(object):
    level_relations = {
        'debug': logging.DEBUG,
        'info': logging.INFO,
        'warning': logging.WARNING,
        'error': logging.ERROR,
        'crit': logging.CRITICAL
    }

    def __init__(self, filename, level='info', when='D', backCount=3,
                 fmt='%(asctime)s - %(pathname)s[line:%(lineno)d] - %(levelname)s: %(message)s'):
        self.logger = logging.getLogger(filename)
        format_str = logging.Formatter(fmt)
        self.logger.setLevel(self.level_relations.get(level))
        sh = logging.StreamHandler()
        sh.setFormatter(format_str)
        th = handlers.TimedRotatingFileHandler(filename=filename, when=when, backupCount=backCount, encoding='utf-8')
        th.setFormatter(format_str)
        self.logger.addHandler(sh)
        self.logger.addHandler(th)


log = Logger('statistics.log', level='debug')


def prepare_env():
    """
    Prepare repository and directory
    """
    log.logger.info('=' * 25 + ' PREPARE ENVIRONMENT ' + '=' * 25)
    if os.path.exists('community'):
        subprocess.call('rm -rf community', shell=True)
    subprocess.call('git clone https://gitee.com/openeuler/community.git', shell=True)
    if not os.path.exists('community'):
        log.logger.error('Fail to clone code, exit...')
        sys.exit(1)
    data_dir = 'data'
    if os.path.exists(data_dir):
        subprocess.call('rm -rf {}'.format(data_dir), shell=True)
    subprocess.call('mkdir data', shell=True)
    if not os.path.exists('data'):
        log.logger.error('Fail to make data directory, exit...')
        sys.exit(1)
    log.logger.info('ENV is already.\n')
    return data_dir


def get_sigs():
    """
    Get relationship between sigs and repositories
    """
    log.logger.info('=' * 25 + ' GET SIGS INFO ' + '=' * 25)
    sig_path = os.path.join('community', 'sig')
    sigs = []
    sigs_list = []
    for i in sorted(os.listdir(sig_path)):
        if i in ['README.md', 'sig-template', 'sig-recycle', 'create_sig_info_template.py']:
            continue
        if i not in [x['name'] for x in sigs]:
            sigs.append({'name': i, 'repositories': []})
            sigs_list.append(i)
        if 'openeuler' in os.listdir(os.path.join(sig_path, i)):
            for filesdir, _, repos in os.walk(os.path.join(sig_path, i, 'openeuler')):
                for repo in repos:
                    for sig in sigs:
                        if sig['name'] == i:
                            repositories = sig['repositories']
                            repositories.append(os.path.join('openeuler', repo.split('.yaml')[0]))
        if 'src-openeuler' in os.listdir(os.path.join(sig_path, i)):
            for filesdir, _, src_repos in os.walk(os.path.join(sig_path, i, 'src-openeuler')):
                for src_repo in src_repos:
                    for sig in sigs:
                        if sig['name'] == i:
                            repositories = sig['repositories']
                            repositories.append(os.path.join('src-openeuler', src_repo.split('.yaml')[0]))
    log.logger.info('Get sigs info.\n')
    return sigs, sigs_list


def get_maintainers(sig):
    """
    Get maintainers of the sig and mark where "maintainers" come from
    :param sig: sig name
    :return: maintainers, sig_info_mark
    """
    owners_file = os.path.join('community', 'sig', sig, 'OWNERS')
    sig_info_file = os.path.join('community', 'sig', sig, 'sig-info.yaml')
    if os.path.exists(owners_file):
        with open(owners_file, 'r', encoding='utf-8') as f:
            maintainers = yaml.load(f.read(), Loader=yaml.Loader)['maintainers']
            return maintainers, False
    elif os.path.exists(sig_info_file):
        with open(sig_info_file, 'r', encoding='utf-8') as f:
            sig_info = yaml.load(f.read(), Loader=yaml.Loader)
            maintainers = [x['gitee_id'] for x in sig_info['maintainers']]
            return maintainers, True
    else:
        log.logger.error('ERROR! Find SIG {} has neither OWNERS file nor sig-info.yaml.'.format(sig))
        sys.exit(1)


def get_committers_mapping(sig):
    """
    Get mappings between repos and committers
    :param sig: sig name
    :return: committers_mapping
    """
    sig_info_file = os.path.join('community', 'sig', sig, 'sig-info.yaml')
    if not os.path.exists(sig_info_file):
        return {}
    with open(sig_info_file, 'r', encoding='utf-8') as f:
        sig_info = yaml.load(f.read(), Loader=yaml.Loader)
    repositories = sig_info.get('repositories')
    if not repositories:
        return {}
    committers_mapping = {}
    for i in repositories:
        if 'committers' in i.keys():
            repos = i['repo']
            committers = [x['gitee_id'] for x in i['committers']]
            for repo in repos:
                committers_mapping[repo] = committers
    return committers_mapping


def get_repo_members(maintainers, committers_mapping, repo):
    """
    Get reviewers of a repo
    :param maintainers: maintainers of the sig
    :param committers_mapping: mappings between repos and committers
    :param repo: full name of repo
    :return: reviewers
    """
    if repo not in committers_mapping.keys():
        return maintainers
    reviewers = committers_mapping[repo]
    return reviewers


def count_duration(start_time):
    """
    Count open days of a Pull Request by its start_time
    :param start_time: time when the Pull Request starts
    :return: duration in days
    """
    today = datetime.datetime.today()
    start_date = datetime.datetime.strptime(start_time, '%Y-%m-%d %H:%M:%S')
    duration = str((today - start_date).days)
    return duration


def create_email_mappings():
    """
    Generate mappings between gitee_id and email addresses
    """
    email_mappings = {}
    if not os.path.exists('community'):
        subprocess.call('git clone https://gitee.com/openeuler/community.git', shell=True)
    sig_path = os.path.join('community', 'sig')
    for i in sorted(os.listdir(sig_path)):
        if i in ['README.md', 'sig-template', 'sig-recycle', 'create_sig_info_template.py']:
            continue
        log.logger.info('Starting to get email mappings of sig {}'.format(i))
        owners_file = os.path.join(sig_path, i, 'OWNERS')
        sig_info_file = os.path.join(sig_path, i, 'sig-info.yaml')
        if os.path.exists(owners_file):
            f = open(owners_file, 'r', encoding='utf-8')
            maintainers = yaml.safe_load(f)['maintainers']
            f.close()
            for maintainer in maintainers:
                if maintainer not in email_mappings.keys():
                    email_mappings[maintainer] = ''
        if os.path.exists(sig_info_file):
            f = open(sig_info_file, 'r', encoding='utf-8')
            sig_info = yaml.safe_load(f)
            f.close()
            maintainers = sig_info['maintainers']
            for maintainer in maintainers:
                maintainer_gitee_id = maintainer['gitee_id']
                maintainer_email = maintainer.get('email')
                if maintainer_email in ['null', 'NA'] or not maintainer_email:
                    maintainer_email = ''
                email_mappings[maintainer_gitee_id] = maintainer_email
            repositories = sig_info.get('repositories')
            if not repositories:
                continue
            for r in repositories:
                if 'committers' in r.keys():
                    commtters = r['committers']
                    for committer in commtters:
                        committer_gitee_id = committer['gitee_id']
                        committer_email = committer.get('email')
                        if committer_email in ['null', 'NA'] or not committer_email:
                            committer_email = ''
                        email_mappings[committer_gitee_id] = committer_email
    ready_to_remove = []
    for email_mapping in email_mappings:
        if not email_mappings[email_mapping]:
            ready_to_remove.append(email_mapping)
    for i in ready_to_remove:
        del email_mappings[i]
    # generate email_mappings.yaml
    with open('email_mapping.yaml', 'w', encoding='utf-8') as f:
        yaml.dump(email_mappings, f, default_flow_style=False)


def get_email_mappings():
    """
    Get email_mappings
    :return: email_mappings
    """
    create_email_mappings()
    if not os.path.exists('email_mapping.yaml'):
        log.logger.error('ERROR! Fail to generate email_mappings.')
        return {}
    email_mappings = yaml.safe_load(open('email_mapping.yaml'))
    return email_mappings


def csv_to_xlsx(filepath):
    """
    Convert a csv file to a xlsx file
    :param filepath: path of the csv file
    :return: path of the xlsx file
    """
    if not filepath.endswith('.csv'):
        return
    # sorting
    df = pd.read_csv(filepath, encoding='utf-8')
    df.to_csv(filepath, mode='w', index=False)

    csv_file = pd.read_csv(filepath, encoding='utf-8')
    xlsx_filepath = filepath.replace('.csv', '.xlsx')
    csv_file.to_excel(xlsx_filepath, sheet_name='open_pull_requests_statistics')
    if not os.path.exists(xlsx_filepath):
        log.logger.error('ERROR! Fail to generate {}'.format(xlsx_filepath))
        sys.exit(1)
    log.logger.info('Generate {}'.format(filepath.replace('.csv', '.xlsx')))
    return xlsx_filepath


def cal_sig_processed_rate(sig_name, ts):
    """
    Calculate processed rate of Pull Requests of a sig between now and a week ago
    :param sig_name: sig name
    :param ts: timestamp
    :return: -1, 0 or a two bit float number
    """
    url = 'https://dsapi.osinfra.cn/query/sig/pr/state'
    params = {
        'community': 'openeuler',
        'timestamp': ts,
        'sig': sig_name
    }
    r = requests.get(url, params=params)
    if r.status_code != 200:
        processed_rate = -1
    else:
        data = r.json()['data']
        if not data:
            return -1
        merged, closed, op = data['merged'], data['closed'], data['open']
        if merged == 0 and closed == 0 and op == 0:
            return 0
        processed_rate = round((merged + closed) / (merged + closed + op), 2)
    return processed_rate


def cal_compare_timestamp():
    """
    Calculate timestamp at 9:00 on the current day and timestamp a week ago
    :return: timestamp at 9:00 on the current day and timestamp a week ago
    """
    from datetime import datetime
    today = datetime.strftime(datetime.today(), '%Y-%m-%d')
    timestamp_today = int(time.mktime(datetime.strptime(today + ' 09', '%Y-%m-%d %H').timetuple())) * 1000
    timestamp_last = timestamp_today - 3600 * 24 * 7 * 1000
    return timestamp_today, timestamp_last


def all_sigs_compare(sigs_list):
    """
    Generate compare info of all sigs
    :param sigs_list: a name list of all sigs
    :return: compare info of all sigs
    """
    compare_dict = {}
    for sig in sigs_list:
        compare_info = compare_sig_processed_rate(sig)
        compare_dict[sig] = compare_info
    return compare_dict


def single_sig_compare(sig, compare_dict):
    """
    Return compare info of a sig
    :param sig: sig name
    :param compare_dict: a dict of every sig and its compare info
    :return: compare info of a sig
    """
    return compare_dict.get(sig)


def compare_sig_processed_rate(sig_name):
    """
    Compare processed rate of a sig
    :param sig_name: sig name
    :return: compare info
    """
    ts_today, ts_last = cal_compare_timestamp()
    processed_rate_now = cal_sig_processed_rate(sig_name, ts_today)
    processed_rate_last = cal_sig_processed_rate(sig_name, ts_last)
    if processed_rate_now == -1 or processed_rate_last == -1:
        return ""
    else:
        if processed_rate_now == processed_rate_last:
            return 'PR处理率为{}%, 同比上周不变'.format(processed_rate_now * 100)
        elif processed_rate_now > processed_rate_last:
            compare_rate = round(processed_rate_now - processed_rate_last, 2)
            return 'PR处理率为{}%, 同比上周上升{}%'.format(processed_rate_now * 100, compare_rate * 100)
        elif processed_rate_now < processed_rate_last:
            compare_rate = round(processed_rate_last - processed_rate_now, 2)
            return 'PR处理率为{}%, 同比上周下降{}%'.format(processed_rate_now * 100, compare_rate * 100)


def excel_optimization(filepath, compare_dict):
    """
    Adjust styles of the xlsx file
    :param filepath: path of the xlsx file
    :param compare_dict: a dict of every sig and its compare info
    """
    if not filepath.endswith('.xlsx'):
        return
    html_file = filepath.replace('.xlsx', '.html')
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    tmp_list = []
    for row in ws.rows:
        tmp_list.append(row[1].value)
    insert_rows = {}
    for i in tmp_list:
        if i not in insert_rows.keys():
            insert_row = tmp_list.index(i) + 1
            insert_rows[insert_row] = i
    # delete auxiliary column
    ws.delete_cols(1)
    ws.delete_cols(1)
    # insert rows
    alignment_center = Alignment(horizontal='center', vertical='center')
    insert_count = 0
    for i in sorted(insert_rows.keys()):
        sig_name = insert_rows[i]
        i += insert_count

        ws.insert_rows(i)
        insert_count += 1
        ws['A' + str(i)] = sig_name
        ws['A' + str(i)].font = Font(name='黑体', size=20, bold=True)
        ws.merge_cells(start_row=i, end_row=i, start_column=1, end_column=6)
        ws['A' + str(i)].alignment = alignment_center

        compare_info = single_sig_compare(sig_name, compare_dict)
        ws.insert_rows(i + 1)
        insert_count += 1
        ws['A' + str(i + 1)] = compare_info
        ws['A' + str(i + 1)].font = Font(name='黑体', color='FF0000')
        ws['A' + str(i + 1)].alignment = alignment_center
        ws.merge_cells(start_row=i + 1, end_row=i + 1, start_column=1, end_column=6)

        ws.insert_rows(i + 2)
        insert_count += 1
        ws['A' + str(i + 2)] = '仓库'
        ws['B' + str(i + 2)] = '目标分支'
        ws['C' + str(i + 2)] = '编号'
        ws['D' + str(i + 2)] = '标题'
        ws['E' + str(i + 2)] = '状态'
        ws['F' + str(i + 2)] = '开启天数'
        ws['A' + str(i + 2)].font = Font(bold=True)
        ws['A' + str(i + 2)].alignment = alignment_center
        ws['B' + str(i + 2)].font = Font(bold=True)
        ws['B' + str(i + 2)].alignment = alignment_center
        ws['C' + str(i + 2)].font = Font(bold=True)
        ws['C' + str(i + 2)].alignment = alignment_center
        ws['D' + str(i + 2)].font = Font(bold=True)
        ws['D' + str(i + 2)].alignment = alignment_center
        ws['E' + str(i + 2)].font = Font(bold=True)
        ws['E' + str(i + 2)].alignment = alignment_center
        ws['F' + str(i + 2)].font = Font(bold=True)
        ws['F' + str(i + 2)].alignment = alignment_center

    # replace the original table header
    ws.insert_rows(5)
    ws['A5'] = ws['A4'].value
    ws['B5'] = ws['B4'].value
    ws['C5'] = ws['C4'].value
    ws['D5'] = ws['D4'].value
    ws['E5'] = ws['E4'].value
    ws['F5'] = ws['F4'].value
    ws.delete_rows(4)
    # fill for the Duration
    cells = ws.iter_rows(min_row=3, min_col=6, max_col=6)
    yellow_fill = PatternFill("solid", start_color='FFFF00')
    first_stage_fill = PatternFill('solid', start_color='FFDAB9')
    second_stage_fill = PatternFill('solid', start_color='FF7F50')
    third_stage_fill = PatternFill('solid', start_color='FF4500')
    for i in cells:
        try:
            value = int(i[0].value)
            if 7 < value <= 30:
                i[0].fill = first_stage_fill
            elif 30 < value <= 365:
                i[0].fill = second_stage_fill
            elif value > 365:
                i[0].fill = third_stage_fill
        except (TypeError, ValueError):
            pass
    # fill for the status mark
    status = ws.iter_rows(min_row=3, min_col=5, max_col=5)
    for j in status:
        value = j[0].value
        if not value:
            continue
        elif len(value) <= 3 and value != '草稿':
            continue
        else:
            j[0].fill = yellow_fill
    # align center
    for row in ws.rows:
        row[5].alignment = alignment_center
    # add borders
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
    for row in ws.rows:
        for cell in row:
            cell.border = border
    wb.save(filepath)
    wb.close()
    # generate html file by the xlsx file
    xlsx2html(filepath, html_file)
    log.logger.info('Generate {}'.format(html_file))


def send_email(xlsx_file, nickname, receivers):
    """
    Send email to reviewers
    :param xlsx_file: path of the xlsx file
    :param nickname: Gitee ID of the receiver
    :param receivers: where send to
    """
    username = os.getenv('SMTP_USERNAME', '')
    port = os.getenv('SMTP_PORT', '')
    host = os.getenv('SMTP_HOST', '')
    password = os.getenv('SMTP_PASSWORD', '')
    msg = MIMEMultipart()
    html_file = xlsx_file.replace('.xlsx', '.html')
    with open(html_file, 'r', encoding='utf-8') as f:
        body_of_email = f.read()
    body_of_email = body_of_email.replace('<body>', '<body><p>Dear {},</p>'
                                                    '<p>以下是您参与openEuler社区的SIG仓库下待处理的PR，烦请您及时跟进</p>'.
                                          format(nickname)).replace('&nbsp;', '0')
    content = MIMEText(body_of_email, 'html', 'utf-8')
    msg.attach(content)
    msg['Subject'] = 'openEuler 待处理PR汇总'
    msg['From'] = username
    msg['To'] = ','.join(receivers)
    try:
        if int(port) == 465:
            server = smtplib.SMTP_SSL(host, port)
            server.ehlo()
            server.login(username, password)
        else:
            server = smtplib.SMTP(host, port)
            server.ehlo()
            server.starttls()
            server.login(username, password)
        server.sendmail(username, receivers, msg.as_string())
        log.logger.info('Sent report email to: {}'.format(receivers))
    except smtplib.SMTPException as e:
        log.logger.error(e)


def fill_status(status, insert_string):
    """
    Change status of the Pull Request
    :param status: a string of current status
    :param insert_string: abnormal status waiting to add
    :return: status
    """
    if status == '待合入':
        status = insert_string
    else:
        status += '、{}'.format(insert_string)
    return status


def clean_env(data_dir):
    """
    Remove the temporary data
    :param data_dir: directory waiting to clean
    """
    subprocess.call('rm -rf {}'.format(data_dir), shell=True)


def get_repos_pulls_mapping():
    """
    Get mappings between repos and pulls
    :return: a dict of {repo: pulls}
    """
    enterprise_pulls = []
    page = 1
    while True:
        log.logger.info("=" * 25 + " GET ENTERPRISE PULLS: PAGE {} ".format(page) + "=" * 25)
        url = 'https://ipb.osinfra.cn/pulls'
        params = {
            'state': 'open',
            'direction': 'asc',
            'page': page,
            'per_page': 100
        }
        r = requests.get(url, params=params)
        if r.status_code != 200:
            log.logger.error('Fail to get enterprise pulls list.')
            return
        else:
            enterprise_pulls += r.json()['data']
        if len(r.json()['data']) < 100:
            break
        page += 1
    return {x['link'].split('/', 3)[3]: x for x in enterprise_pulls}


def pr_statistics(data_dir, sigs, repos_pulls_mapping, compare_dict):
    """
    :param data_dir: directory to store temporary data
    :param sigs: a dict of every sig and its repositories
    :param repos_pulls_mapping: mappings between repos and pulls
    :param compare_dict: a dict of every sig and its compare info
    """
    log.logger.info('=' * 25 + ' STATISTICS ' + '=' * 25)
    email_mappings = get_email_mappings()
    mapping_lists = sorted(list(email_mappings.keys()))
    open_pr_dict = {}
    open_pr_info = []
    for sig in sigs:
        sig_name = sig['name']
        sig_repos = sig['repositories']
        log.logger.info('\nStarting to search sig {}'.format(sig_name))
        if not sig_repos:
            log.logger.info('Find no repositories in sig {}, skip'.format(sig_name))
            continue
        maintainers, sig_info_mark = get_maintainers(sig_name)
        for full_repo in sig_repos:
            if full_repo.split('/')[0] not in ['src-openeuler', 'openeuler']:
                continue
            open_pr_list = []
            for mapping_key in repos_pulls_mapping.keys():
                if mapping_key.startswith(full_repo + '/'):
                    open_pr_list.append(repos_pulls_mapping[mapping_key])
                    log.logger.info('Find open pr: {}'.format(mapping_key))
            if not open_pr_list:
                continue
            members = maintainers
            if sig_info_mark:
                committers_mapping = get_committers_mapping(sig_name)
                members = get_repo_members(maintainers, committers_mapping, full_repo)
            for item in open_pr_list:
                title = item['title']
                html_url = item['link']
                number = '#' + html_url.split('/')[-1]
                created_at = item['created_at']
                draft = item['draft']
                labels = item['labels'].split(',')
                ref_branch = item['ref']
                status = '待合入'
                if draft:
                    status = fill_status(status, '草稿')
                if 'openeuler-cla/yes' not in labels:
                    status = fill_status(status, 'CLA认证失败')
                if 'ci_failed' in labels:
                    status = fill_status(status, '门禁检查失败')
                if not item['mergeable']:
                    status = fill_status(status, '存在冲突')
                if 'kind/wait_for_update' in labels:
                    status = fill_status(status, '等待更新')
                duration = count_duration(created_at)
                link = "<a href='{0}'>{1}</a>".format(html_url, title)
                number_link = "<a href='{0}'>{1}</a>".format(html_url, number)
                open_pr_info.append([sig_name, full_repo, ref_branch, number_link, link, status, duration,
                                     ','.join(members)])
    no_addresses_id = []
    for pr_info in open_pr_info:
        ids = pr_info[-1]
        for i in ids.split(','):
            if i not in mapping_lists:
                if i not in no_addresses_id:
                    log.logger.warning('WARNING! gitee_id {} does not match any email address.'.format(i))
                    no_addresses_id.append(i)
            if i not in open_pr_dict.keys():
                open_pr_dict[i] = [pr_info[:-1]]
            else:
                open_pr_dict[i].append(pr_info[:-1])
    for receiver in sorted(list(open_pr_dict.keys())):
        origin_pr_list = sorted(open_pr_dict[receiver], key=(lambda x: int(x[6])), reverse=True)
        ordered_pr_list = []
        pr_sigs = sorted(set([x[0] for x in origin_pr_list]))
        for pr_sig in pr_sigs:
            for op in origin_pr_list:
                if op[0] == pr_sig:
                    if len(ordered_pr_list) > 0 and op[0] == ordered_pr_list[-1][0] and int(op[-1]) > \
                            int(ordered_pr_list[-1][-1]):
                        ordered_pr_list.insert(-1, op)
                    else:
                        ordered_pr_list.append(op)
        statistics_csv = '{}/statistics_{}.csv'.format(data_dir, receiver)
        f = codecs.open(statistics_csv, 'w', encoding='utf-8')
        writer = csv.writer(f)
        for i in ordered_pr_list:
            writer.writerow(i)
        f.close()
        email_address = email_mappings.get(receiver)
        if not email_address:
            log.logger.warning('Ready to send statistics for {} but cannot find the email address'.format(receiver))
            continue
        log.logger.info('Ready to send statistics for {} whose email address is {}'.format(receiver, email_address))
        statistics_xlsx = csv_to_xlsx(statistics_csv)
        excel_optimization(statistics_xlsx, compare_dict)
        send_email(statistics_xlsx, receiver, [email_address])


def main():
    """
    main function
    """
    data_dir = prepare_env()
    sigs, sigs_list = get_sigs()
    compare_dict = all_sigs_compare(sigs_list)
    print('Compare Dict: {}'.format(compare_dict))
    repos_pulls_mapping = get_repos_pulls_mapping()
    pr_statistics(data_dir, sigs, repos_pulls_mapping, compare_dict)


if __name__ == '__main__':
    main()
